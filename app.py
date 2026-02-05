import os
import sys
import tempfile
import threading
from copy import copy

from flask import Flask, render_template, request, send_file, flash, redirect, jsonify
import openpyxl
from werkzeug.utils import secure_filename


def get_resource_path(relative_path):
    """PyInstallerでパッケージ化された場合のリソースパスを取得"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


app = Flask(__name__, template_folder=get_resource_path('templates'))
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB制限

ALLOWED_EXTENSIONS = {'xlsx', 'xlsm'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_base_filename(filename):
    """拡張子を除いたファイル名を取得"""
    return os.path.splitext(filename)[0]


def merge_workbooks(files_data):
    """
    複数のExcelファイルのシートを1つのワークブックにマージ
    files_data: [(filename, file_object), ...]
    """
    merged_wb = openpyxl.Workbook()
    merged_wb.remove(merged_wb.active)

    sheet_name_counts = {}
    merged_sheets = []

    for original_filename, file_obj in files_data:
        try:
            source_wb = openpyxl.load_workbook(file_obj)
            base_name = get_base_filename(original_filename)

            for sheet_name in source_wb.sheetnames:
                source_sheet = source_wb[sheet_name]

                # シート名の重複を避ける（ファイル名_シート名）
                new_sheet_name = f"{base_name}_{sheet_name}"

                # シート名が31文字を超える場合は切り詰め
                if len(new_sheet_name) > 31:
                    new_sheet_name = new_sheet_name[:28] + "..."

                # 同名シートがある場合は番号を付与
                if new_sheet_name in sheet_name_counts:
                    sheet_name_counts[new_sheet_name] += 1
                    suffix = f"_{sheet_name_counts[new_sheet_name]}"
                    if len(new_sheet_name) + len(suffix) > 31:
                        new_sheet_name = new_sheet_name[:31-len(suffix)] + suffix
                    else:
                        new_sheet_name = new_sheet_name + suffix
                else:
                    sheet_name_counts[new_sheet_name] = 0

                new_sheet = merged_wb.create_sheet(title=new_sheet_name)

                # セルのコピー
                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet[cell.coordinate]
                        new_cell.value = cell.value

                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.alignment = copy(cell.alignment)

                # 列幅のコピー
                for col in source_sheet.column_dimensions:
                    new_sheet.column_dimensions[col].width = \
                        source_sheet.column_dimensions[col].width

                # 結合セルのコピー
                for merged_cell in source_sheet.merged_cells.ranges:
                    new_sheet.merge_cells(str(merged_cell))

                merged_sheets.append({
                    'file': original_filename,
                    'original_sheet': sheet_name,
                    'new_sheet': new_sheet_name
                })

            source_wb.close()

        except Exception as e:
            raise Exception(f"'{original_filename}' の処理中にエラー: {str(e)}")

    return merged_wb, merged_sheets


@app.route('/get_sheets', methods=['POST'])
def get_sheets():
    """アップロードされたExcelファイルのシート名一覧を返す"""
    if 'files[]' not in request.files:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    files = request.files.getlist('files[]')
    result = []

    for file in files:
        if file.filename == '':
            continue

        if not allowed_file(file.filename):
            return jsonify({'error': f'対応していないファイル形式です: {file.filename}'}), 400

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            result.append({
                'filename': file.filename,
                'sheets': sheets,
                'count': len(sheets)
            })
        except Exception as e:
            return jsonify({'error': f'{file.filename} の読み込みに失敗しました: {str(e)}'}), 400

    if not result:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    total_sheets = sum(f['count'] for f in result)
    return jsonify({'files': result, 'total_sheets': total_sheets})


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'files[]' not in request.files:
            flash('ファイルが選択されていません', 'error')
            return redirect(request.url)

        files = request.files.getlist('files[]')
        valid_files = []

        for file in files:
            if file.filename == '':
                continue

            if not allowed_file(file.filename):
                flash(f'対応していないファイル形式です: {file.filename}', 'error')
                return redirect(request.url)

            valid_files.append((file.filename, file))

        if len(valid_files) == 0:
            flash('ファイルが選択されていません', 'error')
            return redirect(request.url)

        if len(valid_files) < 2:
            flash('2つ以上のファイルを選択してください', 'error')
            return redirect(request.url)

        try:
            merged_wb, merged_sheets = merge_workbooks(valid_files)

            temp_dir = tempfile.mkdtemp()
            output_filename = "統合ファイル.xlsx"
            output_path = os.path.join(temp_dir, output_filename)
            merged_wb.save(output_path)

            return send_file(
                output_path,
                as_attachment=True,
                download_name=output_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            flash(f'処理中にエラーが発生しました: {str(e)}', 'error')
            return redirect(request.url)

    return render_template('index.html')


@app.route('/api/merge', methods=['POST'])
def api_merge():
    """Ajax用のマージAPI"""
    if 'files[]' not in request.files:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    files = request.files.getlist('files[]')
    valid_files = []

    for file in files:
        if file.filename == '':
            continue

        if not allowed_file(file.filename):
            return jsonify({'error': f'対応していないファイル形式です: {file.filename}'}), 400

        valid_files.append((file.filename, file))

    if len(valid_files) == 0:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    if len(valid_files) < 2:
        return jsonify({'error': '2つ以上のファイルを選択してください'}), 400

    try:
        merged_wb, merged_sheets = merge_workbooks(valid_files)

        temp_dir = tempfile.mkdtemp()
        output_filename = "統合ファイル.xlsx"
        output_path = os.path.join(temp_dir, output_filename)
        merged_wb.save(output_path)

        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def start_flask_server():
    """Flaskサーバーをバックグラウンドで起動"""
    app.run(host='127.0.0.1', port=5002, debug=False, threaded=True, use_reloader=False)


def main():
    """メインエントリーポイント: pywebviewでデスクトップアプリとして起動"""
    try:
        import webview

        server_thread = threading.Thread(target=start_flask_server, daemon=True)
        server_thread.start()

        import time
        time.sleep(1)

        webview.create_window(
            'SheetMerge - Excelシート統合ツール',
            'http://127.0.0.1:5002',
            width=700,
            height=800,
            resizable=True,
            min_size=(500, 600)
        )
        webview.start()

    except ImportError:
        import webbrowser

        print("pywebviewがインストールされていません。ブラウザで開きます。")
        print("アプリを終了するには Ctrl+C を押してください。")

        def open_browser():
            import time
            time.sleep(1.5)
            webbrowser.open('http://127.0.0.1:5002')

        browser_thread = threading.Thread(target=open_browser, daemon=True)
        browser_thread.start()

        app.run(host='127.0.0.1', port=5002, debug=False, threaded=True)


if __name__ == '__main__':
    main()
